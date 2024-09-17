import telebot
from telebot import types
import logging
import os
from openpyxl import Workbook, load_workbook

# Bot tokeni va admin ID'si
TOKEN = '7328775406:AAGTc8UyUkHoxYaoyXqxgQZJmXR6cplhQFA'
ADMIN_ID = '6760329131'
bot = telebot.TeleBot(TOKEN)

# Excel faylining nomi
EXCEL_FILE = "base.xlsx"
MIN_BALANCE = 5000

# Logger sozlash
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Excel faylini yuklash yoki yaratish
def load_or_create_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["User ID", "Balance"])
        wb.save(EXCEL_FILE)
    else:
        wb = load_workbook(EXCEL_FILE)
    return wb

# Foydalanuvchi hisobini olish
def get_user_balance(user_id):
    wb = load_or_create_excel()
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            return row[1]
    return 0

# Foydalanuvchi hisobini yangilash
def update_user_balance(user_id, amount):
    wb = load_or_create_excel()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == user_id:
            row[1].value = row[1].value + amount
            wb.save(EXCEL_FILE)
            return
    ws.append([user_id, amount])
    wb.save(EXCEL_FILE)

# /start komandasi
@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.chat.id
    balance = get_user_balance(user_id)

    markup = types.InlineKeyboardMarkup()
    mine_button = types.InlineKeyboardButton("⛏ Mayn qilish", callback_data='mine')
    balance_button = types.InlineKeyboardButton("💰 Hisobim", callback_data='balance')
    markup.add(mine_button, balance_button)

    bot.send_message(
        user_id,
        f"👋 Salom! Sizning balansingiz: {balance} so'm.\n💸 Ko'proq topish uchun mayn qilishni boshlang.",
        reply_markup=markup
    )

# Callback funksiyasi
@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    user_id = call.message.chat.id
    if call.data == 'mine':
        update_user_balance(user_id, 2)
        balance = get_user_balance(user_id)
        markup = types.InlineKeyboardMarkup()
        mine_button = types.InlineKeyboardButton("🔁 Yana mayn qilish", callback_data='mine')
        balance_button = types.InlineKeyboardButton("💳 Hisobim", callback_data='balance')
        markup.add(mine_button, balance_button)

        bot.edit_message_text(
            f"🎉 Tabriklaymiz! Siz 2 so'm qo'shdingiz.\n💸 Jami hisobingiz: {balance} so'm.",
            chat_id=user_id,
            message_id=call.message.message_id,
            reply_markup=markup
        )

    elif call.data == 'balance':
        balance = get_user_balance(user_id)
        markup = types.InlineKeyboardMarkup()
        mine_button = types.InlineKeyboardButton("⛏ Mayn qilish", callback_data='mine')
        if balance >= MIN_BALANCE:
            withdraw_button = types.InlineKeyboardButton("💵 Yechib olish", callback_data='withdraw')
            markup.add(mine_button, withdraw_button)
        else:
            markup.add(mine_button)

        bot.edit_message_text(
            f"💸 Jami hisobingiz: {balance} so'm.\n🆔 Id raqamingiz: {user_id}\nMinimal miqdor {MIN_BALANCE} so'm.",
            chat_id=user_id,
            message_id=call.message.message_id,
            reply_markup=markup
        )

# /admin komandasi
@bot.message_handler(commands=['admin'])
def admin(message):
    if str(message.chat.id) != ADMIN_ID:
        bot.send_message(message.chat.id, "Uzr, sizda admin huquqlari yo'q!")
        return
    
    user_list = []
    wb = load_or_create_excel()
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        user_list.append(f"User ID: {row[0]}, Balance: {row[1]} so'm")
    
    if user_list:
        bot.send_message(message.chat.id, "\n".join(user_list))
    else:
        bot.send_message(message.chat.id, "Botda foydalanuvchilar mavjud emas.")

# Botni ishga tushirish
bot.polling()
